# ==========================================
# SIGPET 2.2
# Dashboard Administrativo
# ==========================================

from flask import (
    Blueprint,
    render_template,
    jsonify
)

from database import conectar
from utils import login_requerido

dashboard = Blueprint(
    "dashboard",
    __name__
)

from flask import send_file

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)

from datetime import datetime


# ==========================================
# DASHBOARD
# ==========================================

@dashboard.route("/dashboard")
@login_requerido
def dashboard_admin():

    conexion = conectar()
    cursor = conexion.cursor()

    # ==========================================
    # TOTAL DE PEDIDOS
    # ==========================================

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM pedidos
    """)

    total = cursor.fetchone()["total"]

    # ==========================================
    # PENDIENTES
    # ==========================================

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Pendiente'
    """)

    pendientes = cursor.fetchone()["cantidad"]

    # ==========================================
    # PREPARANDO
    # ==========================================

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Preparando'
    """)

    preparando = cursor.fetchone()["cantidad"]

    # ==========================================
    # LISTOS
    # ==========================================

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Listo'
    """)

    listos = cursor.fetchone()["cantidad"]

    # ==========================================
    # ENTREGADOS
    # ==========================================

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Entregado'
    """)

    entregados = cursor.fetchone()["cantidad"]

    # ==========================================
    # CANCELADOS
    # ==========================================

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Cancelado'
    """)

    cancelados = cursor.fetchone()["cantidad"]

    # ==========================================
    # VENTAS
    # ==========================================

    cursor.execute("""
        SELECT IFNULL(
            SUM(total),
            0
        ) AS ventas

        FROM pedidos

        WHERE estado='Entregado'
    """)

    ventas = cursor.fetchone()["ventas"]

    # ==========================================
    # PROMEDIO
    # ==========================================

    cursor.execute("""
        SELECT IFNULL(
            AVG(total),
            0
        ) AS promedio

        FROM pedidos

        WHERE estado='Entregado'
    """)

    promedio = round(
        cursor.fetchone()["promedio"],
        2
    )

    # ==========================================
    # PRODUCTO FAVORITO
    # ==========================================

    cursor.execute("""
        SELECT

            producto,

            SUM(cantidad) AS total

        FROM detalle_pedido

        GROUP BY producto

        ORDER BY total DESC

        LIMIT 1
    """)

    favorito = cursor.fetchone()

    # ==========================================
    # PRODUCTOS MÁS VENDIDOS
    # ==========================================

    cursor.execute("""
        SELECT

            producto,

            SUM(cantidad) AS total

        FROM detalle_pedido

        GROUP BY producto

        ORDER BY total DESC
    """)

    platillos = []

    for fila in cursor.fetchall():

        platillos.append({

            "platillo": fila["producto"],

            "total": fila["total"]

        })

    # ==========================================
    # VENTAS POR DÍA
    # ==========================================

    cursor.execute("""
        SELECT

            fecha,

            IFNULL(
                SUM(total),
                0
            ) AS ventas

        FROM pedidos

        WHERE estado='Entregado'

        GROUP BY fecha

        ORDER BY fecha
    """)

    ventas_dia = []

    for fila in cursor.fetchall():

        ventas_dia.append({

            "fecha": fila["fecha"],

            "ventas": fila["ventas"]

        })

    conexion.close()



    return render_template(
        "dashboard_admin.html",
        total=total,
        pendientes=pendientes,
        preparando=preparando,
        listos=listos,
        entregados=entregados,
        cancelados=cancelados,
        ventas=ventas,
        promedio=promedio,
        favorito=favorito,
        platillos=platillos,
        ventas_dia=ventas_dia
    )

# ==========================================
# API - ÚLTIMO PEDIDO
# ==========================================

@dashboard.route("/api/ultimo_pedido")
@login_requerido
def ultimo_pedido():

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT
            id,
            cliente,
            estado,
            fecha,
            hora,
            total
        FROM pedidos
        ORDER BY id DESC
        LIMIT 1
    """)

    pedido = cursor.fetchone()

    conexion.close()

    if pedido:

        return jsonify({

            "id": pedido["id"],
            "cliente": pedido["cliente"],
            "estado": pedido["estado"],
            "fecha": pedido["fecha"],
            "hora": pedido["hora"],
            "total": pedido["total"]

        })

    return jsonify({})

# ==========================================
# API DASHBOARD
# ==========================================

@dashboard.route("/api/dashboard")
@login_requerido
def api_dashboard():

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("SELECT COUNT(*) AS total FROM pedidos")
    total = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Pendiente'
    """)
    pendientes = cursor.fetchone()["cantidad"]

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Preparando'
    """)
    preparando = cursor.fetchone()["cantidad"]

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Listo'
    """)
    listos = cursor.fetchone()["cantidad"]

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Entregado'
    """)
    entregados = cursor.fetchone()["cantidad"]

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Cancelado'
    """)
    cancelados = cursor.fetchone()["cantidad"]

    cursor.execute("""
        SELECT IFNULL(SUM(total),0) AS ventas
        FROM pedidos
        WHERE estado='Entregado'
    """)
    ventas = cursor.fetchone()["ventas"]

    cursor.execute("""
        SELECT IFNULL(AVG(total),0) AS promedio
        FROM pedidos
        WHERE estado='Entregado'
    """)
    promedio = cursor.fetchone()["promedio"]

    conexion.close()

    return jsonify({

        "total": total,
        "pendientes": pendientes,
        "preparando": preparando,
        "listos": listos,
        "entregados": entregados,
        "cancelados": cancelados,
        "ventas": ventas,
        "promedio": round(promedio,2)

    })

# ==========================================
# API PLATILLOS
# ==========================================

@dashboard.route("/api/platillos")
@login_requerido
def api_platillos():

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""

        SELECT

            producto,

            SUM(cantidad) AS total

        FROM detalle_pedido

        GROUP BY producto

        ORDER BY total DESC

    """)

    platillos = []

    for fila in cursor.fetchall():

        platillos.append({

            "platillo": fila["producto"],
            "total": fila["total"]

        })

    conexion.close()

    return jsonify(platillos)

# ==========================================
# API VENTAS POR DÍA
# ==========================================

@dashboard.route("/api/ventas_dia")
@login_requerido
def api_ventas_dia():

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""

        SELECT

            fecha,

            IFNULL(SUM(total),0) AS ventas

        FROM pedidos

        WHERE estado='Entregado'

        GROUP BY fecha

        ORDER BY fecha

    """)

    ventas = []

    for fila in cursor.fetchall():

        ventas.append({

            "fecha": fila["fecha"],
            "ventas": fila["ventas"]

        })

    conexion.close()

    return jsonify(ventas)

# ==========================================
# EXPORTAR DASHBOARD EXCEL
# ==========================================

@dashboard.route("/dashboard/excel")
@login_requerido
def dashboard_excel():

    conexion = conectar()
    cursor = conexion.cursor()

    # -------------------------
    # Estadísticas
    # -------------------------

    cursor.execute("SELECT COUNT(*) AS total FROM pedidos")
    total = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Pendiente'
    """)
    pendientes = cursor.fetchone()["cantidad"]

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Preparando'
    """)
    preparando = cursor.fetchone()["cantidad"]

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Listo'
    """)
    listos = cursor.fetchone()["cantidad"]

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Entregado'
    """)
    entregados = cursor.fetchone()["cantidad"]

    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM pedidos
        WHERE estado='Cancelado'
    """)
    cancelados = cursor.fetchone()["cantidad"]

    cursor.execute("""
        SELECT IFNULL(SUM(total),0) AS ventas
        FROM pedidos
        WHERE estado='Entregado'
    """)
    ventas = cursor.fetchone()["ventas"]

    cursor.execute("""
        SELECT IFNULL(AVG(total),0) AS promedio
        FROM pedidos
        WHERE estado='Entregado'
    """)
    promedio = round(cursor.fetchone()["promedio"],2)

    # -------------------------
    # Productos vendidos
    # -------------------------

    cursor.execute("""

        SELECT
            producto,
            SUM(cantidad) AS total

        FROM detalle_pedido

        GROUP BY producto

        ORDER BY total DESC

    """)

    productos = cursor.fetchall()

    conexion.close()

    # =====================================
    # Crear Excel
    # =====================================

    libro = Workbook()

    hoja = libro.active
    hoja.title = "Dashboard"

    azul = PatternFill(
        start_color="0D3B78",
        end_color="0D3B78",
        fill_type="solid"
    )

    blanco = Font(
        color="FFFFFF",
        bold=True
    )

    # -------------------------
    # Encabezado
    # -------------------------

    hoja.merge_cells("A1:B1")

    hoja["A1"] = "Dashboard Administrativo SIGPET"

    hoja["A1"].font = Font(
        bold=True,
        size=18
    )

    hoja["A2"] = "Fecha"

    hoja["B2"] = datetime.now().strftime("%d/%m/%Y %H:%M")

    hoja["A4"] = "Indicador"
    hoja["B4"] = "Valor"

    for celda in hoja["4:4"]:

        celda.fill = azul
        celda.font = blanco
        celda.alignment = Alignment(horizontal="center")

    hoja.append(["Total de pedidos", total])
    hoja.append(["Pendientes", pendientes])
    hoja.append(["Preparando", preparando])
    hoja.append(["Listos", listos])
    hoja.append(["Entregados", entregados])
    hoja.append(["Cancelados", cancelados])
    hoja.append(["Ventas Totales", ventas])
    hoja.append(["Promedio", promedio])

    hoja.column_dimensions["A"].width = 35
    hoja.column_dimensions["B"].width = 25

    # =====================================
    # Segunda hoja
    # =====================================

    hoja2 = libro.create_sheet("Productos")

    hoja2["A1"] = "Producto"
    hoja2["B1"] = "Cantidad"

    for celda in hoja2["1:1"]:

        celda.fill = azul
        celda.font = blanco
        celda.alignment = Alignment(horizontal="center")

    for fila in productos:

        hoja2.append([

            fila["producto"],
            fila["total"]

        ])

    hoja2.column_dimensions["A"].width = 30
    hoja2.column_dimensions["B"].width = 15

    archivo = "Dashboard_SIGPET.xlsx"

    libro.save(archivo)

    return send_file(
        archivo,
        as_attachment=True
    )